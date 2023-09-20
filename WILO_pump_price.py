import tkinter as tk
from tkinter import messagebox
import openpyxl


def save_currency_rate():
    new_rate = currency_entry.get()

    new_rate = new_rate.replace(',', '.')

    if new_rate.replace('.', '', 1).isdigit():
        try:
            with open('currency_rate.txt', 'w') as file:
                file.write(new_rate)
            messagebox.showinfo("Успішно", f"Курс валют оновлено: {new_rate}")
        except Exception as e:
            messagebox.showerror("Помилка", f"Не вдалося зберегти курс валют: {str(e)}")
    else:
        messagebox.showwarning("Увага", "Введіть коректний курс валют (число) перед збереженням!")


def load_currency_rate():
    try:
        with open('currency_rate.txt', 'r') as file:
            rate = file.read()
            if rate:
                currency_entry.delete(0, 'end')
                currency_entry.insert(0, rate)
                return float(rate)
    except Exception as e:
        pass
    return None


def find_pump_price_and_row():
    workbook = None
    price_sheet = None
    pump_price_currency = None
    start_row_index = None

    pump_info = pump_info_entry.get()

    try:
        # Відкриваємо Excel-файл з цінами на насоси
        workbook = openpyxl.load_workbook('price.xlsx')
        price_sheet  = workbook['price_list']
        # Шукаємо відповідний рядок за артикулом або назвою насосу
        for row_num, row in enumerate(price_sheet.iter_rows(values_only=True), start=1):
            print(row)
            if pump_info in row:
                pump_price_currency = row[2]
                start_row_index = row_num
                print(f"pump_price_currency: {pump_price_currency}")
                print(f"start_row_index: {start_row_index}")
                break
    except Exception as e:
        # Обробка можливих помилок при зчитуванні Excel-файлу
        print(f"Помилка при зчитуванні цін: {e}")
    return pump_price_currency, start_row_index, workbook, price_sheet


def find_discount_group_name(price_sheet, start_row_index):
    discount_group = None
    for row in reversed(list(price_sheet.iter_rows(min_row=start_row_index, values_only=True))):
        if "PG" in row:
            discount_group = row[1]
        else:
            discount_group = None
    print(f"Знайдена група знижки: {discount_group}")
    return discount_group


def find_discount_value(workbook, discount_group):
    discount_value = None
    discounts_sheet = workbook['discounts']
    for row in discounts_sheet.iter_rows(values_only=True):
        if discount_group in row:
            discount_value = row[1]
        else:
            discount_value = None
    print(f"Знайдено значення знижки: {discount_value}")
    return discount_value


def calculate_pump_price_uah_and_display():
    workbook, price_sheet, start_row_index, pump_price_currency = find_pump_price_and_row()
    discount_group = find_discount_group_name(price_sheet, start_row_index)
    discount_value = find_discount_value(workbook, discount_group)
    rate = load_currency_rate()
    if pump_price_currency is not None and discount_value is not None:
        pump_price_uah = pump_price_currency * (1 - discount_value/100) * rate
    else:
        pump_price_uah = None
    return pump_price_uah



root = tk.Tk()
root.title("Курс валют")

label = tk.Label(root, text="Курс валют:")
label.pack()

currency_entry = tk.Entry(root)
currency_entry.pack()

currency_save_button = tk.Button(root, text="Зберегти", command=save_currency_rate)
currency_save_button.pack()

# Створення мітки та поля для введення артикулу або назви насосу
pump_info_label = tk.Label(root, text="Назва або артикул насосу:")
pump_info_label.pack()

pump_info_entry = tk.Entry(root)
pump_info_entry.pack()

# Кнопка для зчитування ціни насосу
read_pump_price_button = tk.Button(root, text="Знайти ціну насосу", command=calculate_pump_price_uah_and_display)
read_pump_price_button.pack()

load_currency_rate()

root.mainloop()
